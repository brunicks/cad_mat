from flask import Flask, render_template, request, jsonify, redirect, session, url_for, send_file
import datetime
import os
import requests
import logging
import base64
import json
import sys
import io
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from logging.handlers import RotatingFileHandler
from functools import wraps

# Configuracao simplificada do sistema de logs
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
log_dir = os.path.join(BASE_DIR, 'logs')
os.makedirs(log_dir, exist_ok=True)

# Configurar um único logger para a aplicacao
logger = logging.getLogger('mat_cad')
logger.setLevel(logging.DEBUG)

# Um único formato para todos os logs, com timestamp completo
log_format = logging.Formatter('%(asctime)s [%(levelname)s] [%(name)s:%(lineno)d] [%(process)d] - %(message)s', 
                              datefmt='%Y-%m-%d %H:%M:%S')

# Arquivo de log único com rotacao
file_handler = RotatingFileHandler(
    os.path.join(log_dir, 'application.log'),
    maxBytes=5*1024*1024,  # 5MB
    backupCount=5
)
file_handler.setFormatter(log_format)
file_handler.setLevel(logging.INFO)
logger.addHandler(file_handler)

# Handler para console
console_handler = logging.StreamHandler()
console_handler.setFormatter(log_format)
console_handler.setLevel(logging.DEBUG)
logger.addHandler(console_handler)

# Evitar propagacao para o logger raiz
logger.propagate = False

# Helper simples para serializar objetos JSON para logging
def safe_json_dumps(data):
    try:
        return json.dumps(data, default=lambda o: list(o) if isinstance(o, set) else str(o))
    except:
        return str(data)

# Funcao auxiliar para logs com informacao de usuário
def user_log(level, message):
    """
    Registra logs incluindo informacões do usuário atual quando disponível
    
    Args:
        level: Nível do log (info, debug, warning, error, exception)
        message: Mensagem a ser registrada
    """
    email = session.get('user_email', 'Nao autenticado')
    name = session.get('user_name', '')
    
    # Remover caracteres especiais e acentos do nome
    name = name.encode('ascii', 'ignore').decode('ascii')
    
    # Montar a informacao do usuário
    user_info = f"[{email}"
    if name:
        user_info += f" ({name})]"
    else:
        user_info += "]"
    
    # Remover caracteres especiais e acentos da mensagem
    message = message.encode('ascii', 'ignore').decode('ascii')
    
    full_message = f"{user_info} {message}"
    
    if level == "info":
        logger.info(full_message)
    elif level == "debug":
        logger.debug(full_message)
    elif level == "warning":
        logger.warning(full_message)
    elif level == "error":
        logger.error(full_message)
    elif level == "exception":
        logger.exception(full_message)

# Configuracao da aplicacao Flask
app = Flask(__name__, static_folder='static')
app.secret_key = 'sysmex_material_cadastro_secret_key'

# Constantes da API
SYSMEX_API_BASE_URL = "http://customer-API.qa.sysmexamerica.com/api"
APP_ID = "b51bc34c-52bd-4678-9e08-1580b58c1a79"

# Novas constantes para API de materiais
MATERIAL_API_ENDPOINT = f"{SYSMEX_API_BASE_URL}/Material/AddOrUpdate"
MATERIAL_SEARCH_ENDPOINT = f"{SYSMEX_API_BASE_URL}/Material/GetByFilter"

# Funcao para obter token da aplicacao
def get_app_token():
    """Obtém token da aplicacao via API"""
    try:
        url = f"{SYSMEX_API_BASE_URL}/Usuario/TokenApp"
        headers = {'Content-Type': 'application/json-patch+json'}
        
        # Enviar apenas o app_id entre aspas como corpo da requisicao
        data = f'"{APP_ID}"'
        
        # Menos logs, mais direto
        response = requests.post(url, data=data, headers=headers)
        
        if response.status_code == 200:
            data = response.json()
            if data.get('result') == True:
                return data.get('token')
            else:
                logger.error(f"API recusou token de aplicacao: {safe_json_dumps(data)}")
                return None
        else:
            logger.error(f"Falha na solicitacaoo de token: Status {response.status_code}")
            return None
    except Exception as e:
        logger.exception(f"Excecao ao obter token: {str(e)}")
        return None

# Funcao para autenticar usuário
def authenticate_user(username, password):
    """Autentica usuário com Sysmex API"""
    try:
        # Gerar credenciais em base64
        credentials = f"{username}:{password}"
        credentials_bytes = credentials.encode('ascii')
        base64_credentials = base64.b64encode(credentials_bytes).decode('ascii')
        
        # Obter token da aplicacao
        app_token = get_app_token()
        if not app_token:
            logger.error("Falha ao obter token de aplicacao")
            return None, "Erro ao obter token da aplicacao"
        
        # Request de login
        url = f"{SYSMEX_API_BASE_URL}/Usuario/Login"
        headers = {
            'Authorization': f"Bearer {app_token}",
            'Content-Type': 'application/json-patch+json'
        }
        
        data = f'"{base64_credentials}"'
        
        # Enviar requisicao de autenticacao
        response = requests.post(url, data=data, headers=headers)
        
        if response.status_code == 200:
            user_data = response.json()
            
            # Verificar se o resultado é verdadeiro
            if not user_data.get('result', False):
                logger.error(f"API rejeitou credenciais: {username}")
                return None, "Falha na autenticacao. Verifique suas credenciais."            # Verificar se é um email Sysmex (case-insensitive)
            user_email = user_data.get('usuario', {}).get('email', '')
            if 'sysmex.com' not in user_email.lower():
                logger.warning(f"Acesso bloqueado - email nao Sysmex: {user_email}")
                return None, "Usuario nao autorizado. Apenas emails Sysmex sao permitidos."
            
            user_name = user_data.get('usuario', {}).get('nome', 'Usuário')
            
            # Estrutura correta do objeto de usuário 
            user_info = {
                'token': user_data.get('token'),
                'user': {
                    'email': user_email,
                    'nome': user_name
                }
            }
            
            # Log mais informativo da autenticacao
            logger.info(f"AUTENTICACAO BEM-SUCEDIDA: {user_email} ({user_name})")
            return user_info, None
        else:
            logger.error(f"API retornou erro ({response.status_code}) para usuário: {username}")
            return None, "Credenciais inválidas"
    except Exception as e:
        logger.exception(f"ERRO NA AUTENTICACAO: {username}, Erro: {str(e)}")
        return None, f"Erro de autenticacao: {str(e)}"

# Decorator para rotas protegidas            
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_email' not in session:
            logger.warning(f"ACESSO NEGADO: Tentativa de acesso sem autenticacao à rota {request.path}")
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

# Rotas da aplicacao
@app.route('/')
@login_required
def index():
    user_log("info", "Acesso a pagina inicial")
    return render_template('index.html', user_name=session.get('user_name'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_email' in session:
        # Usar o email já na sessao para o log
        user_log("debug", "Usuario ja autenticado, redirecionando para a pagina inicial")
        return redirect(url_for('index'))
        
    error = None
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if not username or not password:
            logger.warning(f"Tentativa de login com credenciais incompletas: usuário='{username}'")
            error = "Por favor, informe usuário e senha."
            return render_template('login.html', error=error)
        
        # Log antes da autenticacao (nao temos o email Sysmex ainda)
        logger.info(f"Tentativa de login: {username}")
        
        user_data, auth_error = authenticate_user(username, password)
        
        if user_data:
            # Salvar dados do usuário na sessao
            session['user_token'] = user_data['token']
            session['user_name'] = user_data['user'].get('nome', 'Usuário')
            session['user_email'] = user_data['user'].get('email', '')
            
            # Log após autenticacao bem-sucedida
            user_log("info", "LOGIN BEM-SUCEDIDO")
            
            # Redirecionar para a página solicitada ou para a página inicial
            next_page = request.args.get('next', url_for('index'))
            return redirect(next_page)
        else:
            logger.warning(f"Falha no login: usuário='{username}', motivo={auth_error}")
            error = auth_error or "Erro de autenticacao desconhecido."
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    # Capturar email antes de limpar a sessao para o log
    if 'user_email' in session:
        user_log("info", "LOGOUT")
    session.clear()
    return redirect(url_for('login'))

@app.route('/check_auth')
def check_auth():
    if 'user_email' in session:
        return jsonify({'authenticated': True, 'user': session.get('user_name')})
    return jsonify({'authenticated': False})

@app.route('/api/submit_material', methods=['POST'])
@login_required
def submit_material():
    """Envia um material para a API do Sysmex"""
    try:
        material_data = request.json
        
        if not material_data:
            user_log("error", "Tentativa de envio de material sem dados")
            return jsonify({'success': False, 'error': 'Dados do material nao fornecidos'}), 400
        
        # Obter o código do material
        codigo = material_data.get('materialIdExt', '')
        nome = material_data.get('nome', '')
        
        if not codigo:
            user_log("error", "Tentativa de envio de material sem código")
            return jsonify({'success': False, 'error': 'Código do material é obrigatório'}), 400
        
        # Verificar se o material já existe - buscar pelo código
        search_filter = {
            'materialIdExt': codigo
        }
        
        # Configurar cabeçalhos com token de autenticacao
        headers = {
            'Authorization': f"Bearer {session['user_token']}",
            'Content-Type': 'application/json'
        }
        
        # Buscar material existente
        search_response = requests.post(
            MATERIAL_SEARCH_ENDPOINT, 
            json=search_filter, 
            headers=headers
        )
        
        existing_material_id = 0
        
        if search_response.status_code == 200:
            try:
                search_result = search_response.json()
                
                # Verificar se encontrou algum material
                if isinstance(search_result, list):
                    # API retornou lista direta
                    if len(search_result) > 0:
                        # Usar o primeiro material encontrado (deve ser único pelo código)
                        existing_material_id = search_result[0].get('materialId', 0)
                        user_log("info", f"Material com código {codigo} já existe (ID={existing_material_id}). Atualizando.")
                elif search_result.get('result', False) and search_result.get('items'):
                    # API retornou objeto com lista em 'items'
                    items = search_result.get('items', [])
                    if len(items) > 0:
                        existing_material_id = items[0].get('materialId', 0)
                        user_log("info", f"Material com código {codigo} já existe (ID={existing_material_id}). Atualizando.")
            except Exception as e:
                # Se der erro na busca, apenas seguir com cadastro normal
                user_log("warning", f"Erro ao verificar material existente: {str(e)}")
        
        # Completar o payload com valores padrao para a API
        complete_material = {
            "materialId": existing_material_id,  # Se for 0, cria novo; se não, atualiza existente
            "materialIdExt": codigo,
            "nome": nome,
            "tipo": "FERT",  # valor padrao
            "matGrupoId": 7,  # valor padrao
            "uniMedida": "EA",  # valor padrao
            "ativo": True  # valor padrao
        }
        
        # Log com informacões completas do material e usuário
        action = "ATUALIZANDO" if existing_material_id > 0 else "ENVIANDO"
        user_log("info", f"{action} MATERIAL: Código={codigo}, Nome={nome}")
        
        # Configurar cabecalhos com token de autenticacao
        headers = {
            'Authorization': f"Bearer {session['user_token']}",
            'Content-Type': 'application/json'
        }
        
        # Enviar material para a API
        response = requests.post(
            MATERIAL_API_ENDPOINT, 
            json=complete_material, 
            headers=headers
        )
        
        # Analisar resposta
        if response.status_code == 200:
            try:
                response_data = response.json()
                  # Verificar se a API indica sucesso
                if response_data.get('result', False):
                    action = "atualizado" if existing_material_id > 0 else "cadastrado"
                    user_log("info", f"MATERIAL {action.upper()}: Código={codigo}, Nome={nome}")
                    return jsonify({
                        'success': True,
                        'message': f'Material {action} com sucesso',
                        'data': response_data,
                        'wasUpdated': existing_material_id > 0
                    })
                else:
                    error_msg = response_data.get('message', 'Erro nao especificado pela API')
                    user_log("error", f"FALHA NO CADASTRO: Código={codigo}, Erro={error_msg}")
                    return jsonify({
                        'success': False,
                        'error': error_msg,
                        'data': response_data
                    })
            except Exception as e:
                user_log("exception", f"ERRO AO PROCESSAR RESPOSTA: Código={codigo}, Erro={str(e)}")
                return jsonify({
                    'success': False,
                    'error': f"Erro ao processar resposta: {str(e)}"
                }), 500
        else:
            error_msg = f"Erro na comunicacao com a API: {response.status_code}"
            
            try:
                # Tentar obter mensagem de erro do corpo da resposta
                error_data = response.json()
                if error_data and 'message' in error_data:
                    error_msg = error_data['message']
            except:
                # Se falhar ao obter mensagem JSON, usar o texto da resposta
                if response.text:
                    error_msg = f"{error_msg} - {response.text[:200]}"
            
            user_log("error", f"FALHA NO CADASTRO: Código={codigo}, Status={response.status_code}, Erro={error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 500
    except Exception as e:
        user_log("exception", f"EXCEcaO NO CADASTRO: Erro={str(e)}")
        return jsonify({
            'success': False,
            'error': f"Excecao ao processar requisicao: {str(e)}"
        }), 500

@app.route('/material_management')
@login_required
def material_management():
    user_log("info", "Acesso a pagina de gestao de materiais")
    return render_template('material_management.html', user_name=session.get('user_name'))

@app.route('/api/search_materials', methods=['POST'])
@login_required
def search_materials():
    """Pesquisa materiais com base nos filtros fornecidos"""
    try:
        # Obter filtros do request
        filters = request.json or {}
        
        # Criar resumo dos filtros para log
        filtros_resumo = []
        if filters.get('materialIdExt'):
            filtros_resumo.append(f"Código={filters['materialIdExt']}")
        if 'ativo' in filters:
            status = "ativos" if filters['ativo'] else "bloqueados"
            filtros_resumo.append(f"Status={status}")
        
        # Log da pesquisa
        filtros_str = ", ".join(filtros_resumo) if filtros_resumo else "sem filtros"
        page = filters.get('pageNumber', 1)
        page_size = filters.get('pageSize', 25)
        
        user_log("info", f"PESQUISANDO MATERIAIS: {filtros_str} (Página {page}, {page_size} por página)")
        
        # Configurar cabecalhos com token de autenticacao
        headers = {
            'Authorization': f"Bearer {session['user_token']}",
            'Content-Type': 'application/json'
        }
        
        # Enviar pesquisa para a API
        response = requests.post(
            MATERIAL_SEARCH_ENDPOINT, 
            json=filters, 
            headers=headers
        )
        
        # Analisar resposta
        if response.status_code == 200:
            try:
                response_data = response.json()
                
                # Verificar o tipo da resposta - pode ser uma lista ou um objeto
                if isinstance(response_data, list):
                    # API retornou diretamente a lista de materiais
                    items = response_data
                    total_count = len(items)
                    
                    # Se a API nao implementa paginacao, fazemos manualmente
                    start_index = (page - 1) * page_size
                    end_index = start_index + page_size
                    paged_items = items[start_index:end_index] if start_index < len(items) else []
                    
                    # Calcular o número total de páginas
                    total_pages = (total_count + page_size - 1) // page_size if page_size > 0 else 0
                    
                    user_log("info", f"PESQUISA CONCLUÍDA: {len(paged_items)} itens exibidos de um total de {total_count}")
                    
                    return jsonify({
                        'success': True,
                        'data': {
                            'items': paged_items,
                            'totalCount': total_count,
                            'totalPages': total_pages,
                            'pageNumber': page,
                            'pageSize': page_size
                        }
                    })
                else:
                    # Formato esperado como objeto com result, items, etc.
                    if response_data.get('result', False):
                        items = response_data.get('items', [])
                        total_count = response_data.get('count', 0)
                        
                        # Calcular o número total de páginas
                        total_pages = (total_count + page_size - 1) // page_size if page_size > 0 else 0
                        
                        user_log("info", f"PESQUISA CONCLUÍDA: {len(items)} itens exibidos de um total de {total_count}")
                        
                        return jsonify({
                            'success': True,
                            'data': {
                                'items': items,
                                'totalCount': total_count,
                                'totalPages': total_pages,
                                'pageNumber': page,
                                'pageSize': page_size
                            }
                        })
                    else:
                        error_msg = response_data.get('message', 'Erro nao especificado pela API')
                        user_log("error", f"FALHA NA PESQUISA: {filtros_str}, Erro={error_msg}")
                        return jsonify({
                            'success': False,
                            'error': error_msg
                        })
            except Exception as e:
                user_log("exception", f"ERRO AO PROCESSAR PESQUISA: {filtros_str}, Erro={str(e)}")
                return jsonify({
                    'success': False,
                    'error': f"Erro ao processar resposta: {str(e)}"
                }), 500
        else:
            error_msg = f"Erro na comunicacao com a API: {response.status_code}"
            
            try:
                # Tentar obter mensagem de erro do corpo da resposta
                error_data = response.json()
                if isinstance(error_data, dict) and 'message' in error_data:
                    error_msg = error_data['message']
            except:
                # Se falhar ao obter mensagem JSON, usar o texto da resposta
                if response.text:
                    error_msg = f"{error_msg} - {response.text[:200]}"
            
            user_log("error", f"FALHA NA PESQUISA: {filtros_str}, Status={response.status_code}, Erro={error_msg}")
            # Para fins de desenvolvimento, retornar erro 200 em vez de 500
            # para que o cliente possa mostrar a mensagem de erro
            return jsonify({
                'success': False,
                'error': error_msg
            }), 200
    except Exception as e:
        user_log("exception", f"EXCEcaO NA PESQUISA: Erro={str(e)}")
        return jsonify({
            'success': False,
            'error': f"Excecao ao processar requisicao: {str(e)}"
        }), 200  # Código 200 para que o cliente exiba a mensagem

@app.route('/api/update_material', methods=['POST'])
@login_required
def update_material():
    """Atualiza um material existente ou cria um novo"""
    try:
        # Obter dados do material do request
        material_data = request.json
        
        if not material_data:
            user_log("error", "Tentativa de atualizacao de material sem dados")
            return jsonify({'success': False, 'error': 'Dados do material nao fornecidos'}), 400
        
        # Garantir que campos obrigatórios estao preenchidos
        if not material_data.get('materialIdExt') or not material_data.get('nome'):
            user_log("error", "Tentativa de atualizacao com campos obrigatórios faltando")
            return jsonify({
                'success': False, 
                'error': 'Código e nome do material sao obrigatórios'
            }), 400
        
        # Garantir que os campos padrao estao preenchidos
        material_data['tipo'] = material_data.get('tipo', 'FERT')
        material_data['matGrupoId'] = material_data.get('matGrupoId', 7)
        material_data['uniMedida'] = material_data.get('uniMedida', 'EA')
        
        # Informacões importantes para log
        material_id = material_data.get('materialId', 0)
        codigo = material_data.get('materialIdExt')
        nome = material_data.get('nome')
        acao = "ATUALIZANDO" if material_id > 0 else "CADASTRANDO"
        status = "ATIVO" if material_data.get('ativo', True) else "BLOQUEADO"
        
        # Log detalhado da operacao
        user_log("info", f"{acao} MATERIAL: ID={material_id}, Código={codigo}, Nome={nome}, Status={status}")
        
        # Configurar cabecalhos com token de autenticacao
        headers = {
            'Authorization': f"Bearer {session['user_token']}",
            'Content-Type': 'application/json'
        }
        
        # Enviar material para a API
        response = requests.post(
            MATERIAL_API_ENDPOINT, 
            json=material_data, 
            headers=headers
        )
        
        # Analisar resposta
        if response.status_code == 200:
            try:
                response_data = response.json()
                
                # Verificar se a API indica sucesso
                if response_data.get('result', False):
                    action = "ATUALIZADO" if material_id > 0 else "CADASTRADO"
                    user_log("info", f"MATERIAL {action}: ID={material_id}, Código={codigo}, Nome={nome}, Status={status}")
                    return jsonify({
                        'success': True,
                        'message': f'Material {action.lower()} com sucesso',
                        'data': response_data
                    })
                else:
                    error_msg = response_data.get('message', 'Erro nao especificado pela API')
                    user_log("error", f"FALHA NA {acao}: Código={codigo}, Erro={error_msg}")
                    return jsonify({
                        'success': False,
                        'error': error_msg,
                        'data': response_data
                    })
            except Exception as e:
                user_log("exception", f"ERRO AO PROCESSAR RESPOSTA: Código={codigo}, Erro={str(e)}")
                return jsonify({
                    'success': False,
                    'error': f"Erro ao processar resposta: {str(e)}"
                }), 500
        else:
            error_msg = f"Erro na comunicacao com a API: {response.status_code}"
            
            try:
                # Tentar obter mensagem de erro do corpo da resposta
                error_data = response.json()
                if error_data and 'message' in error_data:
                    error_msg = error_data['message']
            except:
                # Se falhar ao obter mensagem JSON, usar o texto da resposta
                if response.text:
                    error_msg = f"{error_msg} - {response.text[:200]}"
            
            user_log("error", f"FALHA NA {acao}: Código={codigo}, Status={response.status_code}, Erro={error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 500
    except Exception as e:
        user_log("exception", f"EXCEcaO NA ATUALIZAcaO: Erro={str(e)}")
        return jsonify({
            'success': False,
            'error': f"Excecao ao processar requisicao: {str(e)}"
        }), 500

@app.route('/api/export_materials', methods=['POST'])
@login_required
def export_materials():
    """Exporta a lista de materiais para um arquivo Excel"""
    try:
        # Obter filtros do request
        filters = request.json or {}
        
        # Criar resumo dos filtros para log
        filtros_resumo = []
        if filters.get('materialIdExt'):
            filtros_resumo.append(f"Código={filters['materialIdExt']}")
        if 'ativo' in filters:
            status = "ativos" if filters['ativo'] else "bloqueados"
            filtros_resumo.append(f"Status={status}")
        
        # Para exportar, queremos todos os resultados, não apenas uma página
        # Removemos paginação e definimos que queremos todos os registros
        export_filters = {k: v for k, v in filters.items() if k not in ['pageNumber', 'pageSize']}
        
        # Log da exportação
        filtros_str = ", ".join(filtros_resumo) if filtros_resumo else "sem filtros"        
        user_log("info", f"EXPORTANDO MATERIAIS PARA EXCEL: {filtros_str}")
        
        # Configurar cabeçalhos com token de autenticação
        headers = {
            'Authorization': f"Bearer {session['user_token']}",
            'Content-Type': 'application/json'
        }
        
        # Enviar pesquisa para a API sem paginação para obter todos os registros
        response = requests.post(
            MATERIAL_SEARCH_ENDPOINT, 
            json=export_filters, 
            headers=headers
        )
        
        # Processar resposta
        if response.status_code != 200:
            error_msg = f"Erro na comunicacao com a API: {response.status_code}"
            
            try:
                # Tentar obter mensagem de erro do corpo da resposta
                error_data = response.json()
                if isinstance(error_data, dict) and 'message' in error_data:
                    error_msg = error_data['message']
            except:
                # Se falhar ao obter mensagem JSON, usar o texto da resposta
                if response.text:
                    error_msg = f"{error_msg} - {response.text[:200]}"
            
            user_log("error", f"FALHA NA EXPORTACAO: Status={response.status_code}, Erro={error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 500
        
        # Analisar dados da resposta
        try:
            response_data = response.json()
            
            # Verificar o tipo da resposta - pode ser uma lista ou um objeto
            if isinstance(response_data, list):
                # API retornou diretamente a lista de materiais
                materials = response_data
            else:
                # Formato esperado como objeto com result, items, etc.
                if response_data.get('result', False):
                    materials = response_data.get('items', [])
                else:
                    error_msg = response_data.get('message', 'Erro nao especificado pela API')
                    user_log("error", f"FALHA NA EXPORTACAO: {filtros_str}, Erro={error_msg}")
                    return jsonify({
                        'success': False,
                        'error': error_msg
                    })
        except Exception as e:
            user_log("exception", f"ERRO AO PROCESSAR DADOS PARA EXPORTACAO: {filtros_str}, Erro={str(e)}")
            return jsonify({
                'success': False,
                'error': f"Erro ao processar resposta: {str(e)}"
            }), 500
        
        # Verificar se há dados para exportar
        if not materials:
            user_log("warning", f"EXPORTACAO VAZIA: {filtros_str}")
            return jsonify({
                'success': False,
                'error': "Nenhum material encontrado para exportar"
            }), 404
        
        # Verificar se há muitos registros (limite para evitar sobrecarga)
        MAX_EXPORT = 10000  # Limite de registros para exportacao
        total_materials = len(materials)
        is_truncated = False
        
        if total_materials > MAX_EXPORT:
            user_log("warning", f"EXPORTACAO MUITO GRANDE: {total_materials} registros. Limitando a {MAX_EXPORT}.")
            materials = materials[:MAX_EXPORT]  # Limitar aos primeiros MAX_EXPORT registros
            is_truncated = True
        
        # Criar arquivo Excel em memória
        try:
            # Criar workbook e selecionar a planilha ativa
            wb = Workbook()
            ws = wb.active
            ws.title = "Materiais"
            
            # Definir estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0056B3", end_color="0056B3", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            warning_font = Font(bold=True, color="FF0000")
            
            row_num = 1
            
            # Adicionar informação sobre limitação de registros se necessário
            if is_truncated:
                warning_text = f"* ATENCAO: Exportacao limitada a {MAX_EXPORT} de {total_materials} registros para evitar sobrecarga."
                warning_cell = ws.cell(row=row_num, column=1)
                warning_cell.value = warning_text
                warning_cell.font = warning_font
                ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
                row_num += 1
            
            # Cabeçalhos
            headers = ["Código", "Nome/Descrição", "Status", "Tipo", "Unidade"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Adicionar dados - começando da próxima linha após os cabeçalhos
            row_num += 1  # Incrementa para a primeira linha de dados
            
            for material in materials:
                # Código
                cell = ws.cell(row=row_num, column=1)
                cell.value = material.get('materialIdExt', '')
                cell.border = thin_border
                
                # Nome/Descrição
                cell = ws.cell(row=row_num, column=2)
                cell.value = material.get('nome', '')
                cell.border = thin_border
                
                # Status
                cell = ws.cell(row=row_num, column=3)
                cell.value = "Ativo" if material.get('ativo', True) else "Bloqueado"
                cell.border = thin_border
                
                # Tipo
                cell = ws.cell(row=row_num, column=4)
                cell.value = material.get('tipo', 'FERT')
                cell.border = thin_border
                
                # Unidade
                cell = ws.cell(row=row_num, column=5)
                cell.value = material.get('uniMedida', 'EA')
                cell.border = thin_border
                
                row_num += 1
            
            # Ajustar largura das colunas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Obter letra da coluna
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # Salvar o arquivo em um buffer temporário
            data_timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            # Gerar nome do arquivo com timestamp
            filename = f"materiais_{data_timestamp}.xlsx"
            
            # Log do sucesso da exportacao
            if is_truncated:
                user_log("info", f"EXPORTACAO CONCLUIDA COM TRUNCAMENTO: {len(materials)} de {total_materials} materiais exportados")
            else:
                user_log("info", f"EXPORTACAO CONCLUIDA: {len(materials)} materiais exportados")
            
            # Enviar o arquivo como resposta
            # Esta abordagem envia o arquivo diretamente para download
            return send_file(
                excel_file,
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        except Exception as e:
            user_log("exception", f"ERRO AO GERAR ARQUIVO EXCEL: {str(e)}")
            return jsonify({
                'success': False,
                'error': f"Erro ao gerar arquivo Excel: {str(e)}"
            }), 500
    except Exception as e:
        user_log("exception", f"EXCECAO NA EXPORTACAO: Erro={str(e)}")
        return jsonify({
            'success': False,
            'error': f"Excecao ao processar requisicao: {str(e)}"
        }), 500

@app.errorhandler(404)
def page_not_found(e):    # Se houver usuário logado, usa user_log, senao usa logger direto
    if 'user_email' in session:
        user_log("warning", f"Pagina nao encontrada: {request.path}")
    else:
        logger.warning(f"Pagina nao encontrada: {request.path}")
    return render_template('login.html', error="Pagina nao encontrada"), 404

@app.errorhandler(500)
def server_error(e):
    # Se houver usuário logado, usa user_log, senao usa logger direto
    if 'user_email' in session:
        user_log("error", f"Erro interno do servidor: {str(e)}")
    else:
        logger.error(f"Erro interno do servidor: {str(e)}")
    return render_template('login.html', error="Erro interno do servidor"), 500

if __name__ == '__main__':
    # Garantir que os diretórios existam
    os.makedirs('templates', exist_ok=True)
    os.makedirs('static', exist_ok=True)
    
    # Adicionar log de inicializacao da aplicacao com data e hora
    start_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    logger.info(f"=== INICIANDO SISTEMA DE CADASTRO DE MATERIAIS v1.0 - {start_time} ===")
    logger.info(f"=== SERVIDOR INICIADO: 0.0.0.0:5000, PID={os.getpid()} ===")
    
    app.run(host='0.0.0.0', port=5000, debug=False)
